# -*- coding:utf-8 -*-

import math
from collections import Counter
from openpyxl import load_workbook, Workbook
from stop_words import get_stop_words
from operator import itemgetter
from numpy import *

class Dataset:
	def __init__(self, listArray, stopWords, minimum_word_length = 3, minimum_word_count = 3):
		self.dataset = {}
		self.dataLine = {}
		self.wordList = []
		self.minLength = minimum_word_length
		self.minCount = minimum_word_count
		self.wordCount = {}
		self.loadData(listArray, stopWords)

	def __str__(self):
		return repr([self.dataset, self.dataLine])

	def loadData(self, lines, stopWords):
		stopWords.extend(get_stop_words("en"))
		stopWords = set(stopWords)
		itab = '!"#$%&()*+,./:;<=>?[\]^_`{|}~-@\''
		otab = '                                '
		# itab = '!",.:;?#$%|{}[]()-'
		# otab = '                  '
	
		buffer = []
		for line in lines:
			key = line[:9]
			value = line[12:]
			# values = set(value.translate(str.maketrans(itab,otab)).lower().split())
			values = value.translate(str.maketrans(itab,otab)).lower().split()
			values = [x for x in values if len(x) >= self.minLength and not x.isdigit()]	
			for stopWord in stopWords:
				while values.count(stopWord) > 0 : values.remove(stopWord)
			self.dataset[key] = values
			self.dataLine[key] = value
			buffer.extend(values)

		self.wordCount = Counter(buffer)
		for k, v in self.wordCount.items():
			if v >= self.minCount: self.wordList.append(k)

def readFile(filename):
	with open(filename, 'rt',encoding = "ISO-8859-1") as f:
		lines = f.readlines()
	return lines		

def loadExcel(filename, worksheet=None):
	wb = load_workbook(filename)
	sheetNames = []; dic = {}
	if worksheet is None: sheetNames = wb.get_sheet_names()
	else: sheetNames.append(worksheet)
	for name in sheetNames:
		ws = wb[name]
		data = []
		for row in ws.rows:
			for cell in row:
				if cell.value != None: data.append(str(cell.value).strip())
		dic[name] = data
	return dic

def dict2array(dic, header):
	fmt = ['f8' for i in range(len(header))]
	data_type = dict(names = header, formats = fmt)
	return array([[key,val] for (key,val) in array.items()], data_type)

def dict2list(dic):
	result = []
	for key, values in dic.items():
		if type(values) != list:
			result += [[key] + [values]]
		else:
			result += [[key] + values]
	return result

def reportResult(dic, filename="Dataset.xlsx"):
	wb = Workbook()
	for sheetName, dataset in dic.items():
		ws = wb.create_sheet(title=sheetName)
		for data in dataset:
			ws.append(data)
	try:
		wb.save(filename)
	except PermissionError as e:
		print(e)
